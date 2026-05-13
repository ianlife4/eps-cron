"""
Excel 報告產生器
產出多分頁 Excel：
  摘要 / 今日新公告 / 完整 EPS / 月營收 / 排行榜
"""
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 樣式 — 對齊 v3 視覺 (Office Word/Excel 預設配色)
FONT_TITLE = Font(name='Microsoft JhengHei', bold=True, color='FFFFFF', size=16)
FONT_TITLE_MD = Font(name='Microsoft JhengHei', bold=True, color='FFFFFF', size=13)
FONT_HEADER = Font(name='Microsoft JhengHei', bold=True, color='FFFFFF', size=11)
FONT_BODY = Font(name='Microsoft JhengHei', size=10)
FONT_HIGHLIGHT = Font(name='Microsoft JhengHei', bold=True, color='9C0006', size=10)

# 標題色塊 — 對齊 v3
FILL_TITLE_BLUE = PatternFill('solid', start_color='2F5496')   # 通用標題 / 高 EPS 排行
FILL_TITLE_RED = PatternFill('solid', start_color='C00000')    # 贏全年確認 (好消息強調)
FILL_TITLE_GOLD = PatternFill('solid', start_color='BF8F00')   # 候選名單 (待驗證)
FILL_TITLE_GREEN = PatternFill('solid', start_color='548235')  # 摘要

FILL_HEADER = PatternFill('solid', start_color='2F5496')       # 欄位 header
FILL_HIGHLIGHT = PatternFill('solid', start_color='FFD966')    # 鮮黃 — 重點列 (對齊 v3)
FILL_HOT = PatternFill('solid', start_color='F8CBAD')          # 暖橘 — 次強調
FILL_COLD = PatternFill('solid', start_color='B4C7E7')         # 冷藍 — 衰退警示
FILL_TODAY = PatternFill('solid', start_color='98F098')        # 鮮綠 — 今日剛公告 (蓋過 score 色)
FILL_BANNER = PatternFill('solid', start_color='DEEBF7')       # 淡藍底 — 區塊標題

FONT_TODAY = Font(name='Microsoft JhengHei', bold=True, color='006600', size=10)

ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
ALIGN_LEFT = Alignment(horizontal='left', vertical='center', indent=1)
ALIGN_RIGHT = Alignment(horizontal='right', vertical='center')
THIN = Side(border_style='thin', color='BFBFBF')
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
BORDER_THICK = Border(top=Side(border_style='medium', color='2F5496'),
                       bottom=Side(border_style='medium', color='2F5496'))


def _style_header_row(ws, row: int, cols: int):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER


def _set_widths(ws, widths: list):
    """list of widths for col A, B, C, ... 對齊 v3 客製寬度。"""
    for i, w in enumerate(widths, 1):
        if w:
            ws.column_dimensions[get_column_letter(i)].width = w


def _draw_title(ws, row: int, text: str, span_cols: int, fill: PatternFill,
                big: bool = True, align: str = 'center'):
    """畫色塊大標題 — 對齊 v3 風格 (預設置中)。"""
    # 整列填底色
    for c in range(1, span_cols + 1):
        ws.cell(row=row, column=c).fill = fill
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = FONT_TITLE if big else FONT_TITLE_MD
    cell.fill = fill
    cell.alignment = Alignment(horizontal=align, vertical='center')
    if span_cols > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols)
    ws.row_dimensions[row].height = 32 if big else 24


def _format_reasons(reasons, max_items: int = 3) -> str:
    """規則版 reasons 是 list[str]，AI 版是 str。統一輸出成 '、' 連接的字串。"""
    if not reasons:
        return ''
    if isinstance(reasons, str):
        return reasons
    if isinstance(reasons, list):
        return '、'.join(str(x) for x in reasons[:max_items])
    return str(reasons)


def _autofit(ws, min_width: int = 8, max_width: int = 30):
    for col_idx, col in enumerate(ws.columns, 1):
        max_len = min_width
        for cell in col:
            try:
                v = str(cell.value) if cell.value is not None else ''
                # 中文佔 2 寬度
                w = sum(2 if ord(ch) > 127 else 1 for ch in v)
                max_len = max(max_len, w)
            except Exception:
                pass
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, max_width)


def _prior_year(q_label: str) -> int:
    """'2026Q1' -> 2025"""
    return int(q_label[:4]) - 1


def _is_won_full_year(r: dict) -> bool:
    """是否確認本期累計已贏去年全年（含虧轉盈）"""
    ach = r.get('achievement_pct')
    if isinstance(ach, (int, float)) and ach >= 1.0:
        return True
    if ach == 'prior_loss' and (r.get('latest_eps') or 0) > 0:
        return True
    pf = r.get('prior_year_full')
    if pf is not None and pf < 0 and (r.get('latest_eps') or 0) > 0:
        return True
    return False


def write_summary(ws, date_str: str, stats: dict, releases: list = None, q_label: str = None,
                  has_revenue: bool = True):
    """摘要分頁 — 對齊 v3「2026Q1 EPS 公告 vs 2025 比較整理」layout:
      [藍底大標 + 副標 (置中)]
      [紅 banner: 已確認 EPS 已贏全年 (置中)]
      [Header + 黃高亮資料表]
      [紅 banner: 重要觀察 (置中)]
      [bullet 表格列]
      [紅 banner: 工作表索引 (置中)]
      [Sheet 名 + 描述 表格]
      [灰斜字 footer]
    """
    _set_widths(ws, [11, 14, 26, 14, 8, 55])
    prior_year = _prior_year(q_label) if q_label else None

    # === [1] 大藍標題 + 副標 (置中) ===
    if q_label:
        title = f'{q_label} EPS 日報 vs {prior_year} 比較整理'
    else:
        title = f'EPS 日報 {date_str}'
    _draw_title(ws, 1, title, 6, FILL_TITLE_BLUE, big=True, align='center')

    for c in range(1, 7):
        ws.cell(row=2, column=c).fill = FILL_TITLE_BLUE
    sub_text = (f'資料來源：FinMind API ({stats.get("total_count", 0)} 檔分析、'
                f'{stats.get("new_count", 0)} 檔今日新公告)　|　日期：{date_str}　|　'
                f'AI 評分 by Claude Haiku 4.5')
    cell = ws.cell(row=2, column=1, value=sub_text)
    cell.font = Font(name='Microsoft JhengHei', italic=True, size=10, color='FFFFFF')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
    ws.row_dimensions[2].height = 22

    row = 4

    # === [2] 已確認已贏全年 (紅 banner + 黃高亮表) ===
    if releases and q_label:
        won = [r for r in releases
               if r.get('latest_quarter') == q_label and _is_won_full_year(r)]
        won.sort(key=lambda x: -(x.get('latest_eps') or 0))

        if won:
            _draw_title(ws, row,
                        f'🏆 已確認 {q_label} EPS 已贏 {prior_year} 全年 ({len(won)} 檔)',
                        6, FILL_TITLE_RED, big=False, align='center')
            row += 1
            sub_headers = ['代號', '名稱', f'{q_label} / {prior_year}全年', '倍數', '評分', '重點']
            for c, h in enumerate(sub_headers, 1):
                cell = ws.cell(row=row, column=c, value=h)
                cell.font = FONT_HEADER
                cell.fill = FILL_HEADER
                cell.alignment = ALIGN_CENTER
                cell.border = BORDER
            ws.row_dimensions[row].height = 22
            row += 1
            for r in won[:10]:
                ach = r.get('achievement_pct')
                pf = r.get('prior_year_full')
                eps = r.get('latest_eps')
                if isinstance(ach, (int, float)):
                    ratio = f'{eps} / {pf} = {ach:.2f}x'
                    multi = f'{ach:.2f}x'
                else:
                    ratio = f'{eps} / {pf} (虧轉盈)' if pf is not None else f'{eps} (虧轉盈)'
                    multi = '虧轉盈'
                reasons = _format_reasons(r.get('reasons'), max_items=2)
                cells = [
                    (1, r.get('stock_id'), ALIGN_CENTER),
                    (2, r.get('name', ''), ALIGN_CENTER),
                    (3, ratio, ALIGN_CENTER),
                    (4, multi, ALIGN_CENTER),
                    (5, r.get('score'), ALIGN_CENTER),
                    (6, reasons, ALIGN_LEFT),
                ]
                for col, val, align in cells:
                    cell = ws.cell(row=row, column=col, value=val)
                    cell.fill = FILL_HIGHLIGHT
                    cell.font = FONT_HIGHLIGHT
                    cell.alignment = align
                    cell.border = BORDER
                ws.row_dimensions[row].height = 22
                row += 1
            row += 1

        # === [3] 重要觀察 (紅 banner + 表格列) ===
        cur_q = [r for r in releases if r.get('latest_quarter') == q_label]
        score_dist = {}
        for r in cur_q:
            s = r.get('score') or 0
            if s >= 6:
                score_dist[s] = score_dist.get(s, 0) + 1
        top3 = sorted([r for r in cur_q if r.get('latest_eps') is not None],
                      key=lambda x: -x['latest_eps'])[:3]
        cand_count = sum(1 for r in cur_q
                         if r.get('prev_quarter_eps') is not None and r['prev_quarter_eps'] < 0
                         and (r.get('latest_eps') or 0) > 0
                         and not _is_won_full_year(r))

        _draw_title(ws, row, '📊 重要觀察', 6, FILL_TITLE_RED, big=False, align='center')
        row += 1

        bullets = []
        if score_dist:
            dist_str = '、'.join(f'+{s} 分 {n} 檔' for s, n in sorted(score_dist.items(), reverse=True))
            bullets.append(f'共 {len(cur_q)} 檔有完整 EPS 資料；高分分布：{dist_str}')
        else:
            bullets.append(f'共 {len(cur_q)} 檔有完整 EPS 資料')
        if top3:
            tstr = '、'.join(f'{r.get("name") or r["stock_id"]} {r["latest_eps"]}' for r in top3)
            bullets.append(f'{q_label} EPS 最高三名：{tstr}')
        if won:
            won_names = '、'.join(f'{r.get("name") or r["stock_id"]}' for r in won[:5])
            bullets.append(f'已確認 {q_label} 一季賺贏 {prior_year} 全年 {len(won)} 檔：{won_names}{"…" if len(won) > 5 else ""}')
        non_won_high = [r for r in cur_q if (r.get('score') or 0) >= 7 and not _is_won_full_year(r)]
        if non_won_high:
            non_won_high.sort(key=lambda x: -(x.get('latest_eps') or 0))
            ach_list = []
            for r in non_won_high[:5]:
                ach = r.get('achievement_pct')
                if isinstance(ach, (int, float)):
                    ach_list.append(f'{r.get("name") or r["stock_id"]} ({ach*100:.0f}%)')
            if ach_list:
                bullets.append(f'高分但尚未贏全年（達成率參考）：{"、".join(ach_list)}')
        if cand_count:
            bullets.append(f'候選名單（上季虧損 + {q_label} 轉正，待全年驗證）：{cand_count} 檔（見 Sheet 「{q_label}贏全年候選」）')

        for b in bullets:
            cell = ws.cell(row=row, column=1, value=b)
            cell.font = Font(name='Microsoft JhengHei', size=11, color='1F3864')
            cell.alignment = Alignment(horizontal='left', vertical='center', indent=1, wrap_text=True)
            for c in range(1, 7):
                ws.cell(row=row, column=c).border = BORDER
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            ws.row_dimensions[row].height = 24
            row += 1
        row += 1

    # === [4] 工作表索引 (紅 banner + 表格) ===
    _draw_title(ws, row, '📑 工作表索引', 6, FILL_TITLE_RED, big=False, align='center')
    row += 1

    sheets_info = [
        ('摘要', '本頁 — 重點摘要 + 已贏全年清單 + 觀察 + 索引'),
        ('今日新公告', f'{stats.get("new_count", 0)} 檔今日新公告（依分數排序）'),
    ]
    if q_label and releases:
        won_n = len([r for r in releases if r.get('latest_quarter') == q_label and _is_won_full_year(r)])
        cand_n = sum(1 for r in releases
                     if r.get('latest_quarter') == q_label
                     and r.get('prev_quarter_eps') is not None and r['prev_quarter_eps'] < 0
                     and (r.get('latest_eps') or 0) > 0
                     and not _is_won_full_year(r))
        top_n = len([r for r in releases if r.get('latest_quarter') == q_label and r.get('latest_eps') is not None])
        if won_n:
            sheets_info.append((f'{q_label}贏全年確認', f'已驗證 {q_label} 賺贏 {prior_year} 全年（{won_n} 檔）'))
        if cand_n:
            sheets_info.append((f'{q_label}贏全年候選', f'上季虧損 + {q_label} 轉正（{cand_n} 檔，待全年驗證）'))
        if top_n:
            sheets_info.append((f'{q_label}高EPS排行', f'{q_label} EPS 前 30 名'))

    hi_n = stats.get('hot_count', 0) + stats.get('watch_count', 0)
    sheets_info.append(('高分排行', f'評分 ≥ 4（{hi_n} 檔）'))
    if stats.get('warn_count', 0) > 0:
        sheets_info.append(('衰退警示', f'評分 ≤ -4（{stats.get("warn_count", 0)} 檔）'))
    sheets_info.append(('完整 EPS', f'全市場 {stats.get("total_count", 0)} 檔依分數排序'))
    if has_revenue:
        sheets_info.append(('月營收', '最新月營收 + YoY + 累計'))

    # Sheet 名跨 A:B (寬 11+14 = 25 足夠 "2026Q1贏全年確認"),描述跨 C:F
    for sheet_name, desc in sheets_info:
        c1 = ws.cell(row=row, column=1, value=sheet_name)
        c1.font = Font(name='Microsoft JhengHei', bold=True, size=11, color='1F3864')
        c1.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        c1.border = BORDER
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.cell(row=row, column=2).border = BORDER

        c2 = ws.cell(row=row, column=3, value=desc)
        c2.font = Font(name='Microsoft JhengHei', size=10, color='000000')
        c2.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        c2.border = BORDER
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
        for c in range(4, 7):
            ws.cell(row=row, column=c).border = BORDER
        ws.row_dimensions[row].height = 22
        row += 1
    row += 1

    # === [5] Footer (灰斜字) ===
    foot = (f'註：所有 EPS / 營收資料來自 FinMind API；驚喜度評分由 Anthropic Claude Haiku 4.5 '
            f'依「驚喜度 / 延續性 / 獲利品質」三維度產出（範圍 -9 ~ +9）。'
            f'達成率 = 今年累計 EPS / 去年全年 EPS。「上季 EPS」指最近一季之前那季'
            f'（例如本期 Q1 → 上季 = 去年 Q4）。'
            f'產出時間：{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}')
    cell = ws.cell(row=row, column=1, value=foot)
    cell.font = Font(name='Microsoft JhengHei', italic=True, size=9, color='808080')
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=1)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.row_dimensions[row].height = 60


def write_releases(ws, releases: list, title: str = '今日新公告', title_fill: PatternFill = None,
                   first_seen_map: dict = None, today_str: str = None):
    """新公告 / 完整 EPS / 高分 / 衰退 通用分頁 — 對齊 v3 視覺。

    若有 first_seen_map → 加「公告日期」欄, 同日期不同色帶。
    今日剛公告 (first_seen == today_str) → 整列鮮綠, 蓋過 score 色, 最醒目。
    依分數降冪後, 同分內以 latest_eps 降冪做 tiebreak。
    """
    has_date = bool(first_seen_map)
    headers = ['代號', '名稱', '市場', '最新季', 'EPS', '去年同季', 'YoY %', 'YoY Δ',
               '累計 EPS', '去年全年', '達成率', 'QoQ %',
               '評分', '級別', '評分理由', 'GM%', 'OPM%', '業外%']
    widths = [9, 14, 8, 10, 9, 11, 11, 11, 11, 11, 11, 11, 7, 14, 36, 9, 9, 9]
    if has_date:
        headers.append('公告日期')
        widths.append(13)
    # v3 風格欄寬 (代號窄 / 名稱中 / 數值 11~13 / 理由 30+)
    _set_widths(ws, widths)

    # 大標題色塊
    _draw_title(ws, 1, title, len(headers), title_fill or FILL_TITLE_BLUE, big=True)

    # Header
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER
    ws.row_dimensions[2].height = 22

    # 二次保險: 評分 desc, latest_eps desc
    sorted_rows = sorted(releases,
                         key=lambda x: (-(x.get('score') if x.get('score') is not None else -99),
                                        -(x.get('latest_eps') or 0)))

    for r in sorted_rows:
        yoy = r.get('yoy') or {}
        qoq = r.get('qoq') or {}
        accum = r.get('accumulated') or {}
        ach = r.get('achievement_pct')
        ach_str = (f'{ach*100:.1f}%' if isinstance(ach, (int, float)) else
                   ('去年虧損' if ach == 'prior_loss' else '—'))
        row_data = [
            r.get('stock_id'),
            r.get('name', ''),
            r.get('market', ''),
            r.get('latest_quarter', ''),
            r.get('latest_eps'),
            r.get('yoy_eps'),
            f'{yoy.get("pct", 0)*100:+.1f}%' if yoy.get('pct') is not None else '—',
            yoy.get('delta'),
            accum.get('value'),
            r.get('prior_year_full'),
            ach_str,
            f'{qoq.get("pct", 0)*100:+.1f}%' if qoq.get('pct') is not None else '—',
            r.get('score'),
            r.get('level', ''),
            _format_reasons(r.get('reasons'), max_items=3),
            f'{(r.get("gm_pct") or 0)*100:.1f}%' if r.get('gm_pct') is not None else '',
            f'{(r.get("opm_pct") or 0)*100:.1f}%' if r.get('opm_pct') is not None else '',
            f'{(r.get("nonop_pct") or 0)*100:.1f}%' if r.get('nonop_pct') is not None else '',
        ]
        if has_date:
            sid = r.get('stock_id')
            row_data.append(first_seen_map.get(sid, '—'))
        ws.append(row_data)

    # 同日色帶 mapping (給 公告日期 欄專用底色)
    date_band_palette = [
        PatternFill('solid', start_color='C6EFCE'),  # 淡綠 (最新)
        PatternFill('solid', start_color='FFEB9C'),  # 淡黃
        PatternFill('solid', start_color='FCE4D6'),  # 淡橘
        PatternFill('solid', start_color='DDEBF7'),  # 淡藍
        PatternFill('solid', start_color='F4E1F2'),  # 淡紫
        PatternFill('solid', start_color='D9D9D9'),  # 淺灰
    ]
    # 推導「今日」: 優先 today_str 參數, 否則 first_seen_map 最大值
    today_for_highlight = today_str
    if not today_for_highlight and first_seen_map:
        valid = [d for d in first_seen_map.values() if d and d != '—']
        if valid:
            today_for_highlight = max(valid)

    date_to_band = {}
    if has_date:
        unique_dates = sorted({first_seen_map.get(r.get('stock_id'), '—') for r in sorted_rows},
                              reverse=True)
        for i, d in enumerate(unique_dates):
            date_to_band[d] = date_band_palette[i % len(date_band_palette)]

    # 上色 + body 樣式 + alignment
    date_col_idx = len(headers) if has_date else None
    for row_idx in range(3, ws.max_row + 1):
        score = ws.cell(row=row_idx, column=13).value
        # 公告日期 value (放在最後欄)
        first_date = ws.cell(row=row_idx, column=date_col_idx).value if date_col_idx else None
        is_today = (first_date == today_for_highlight) if today_for_highlight else False

        fill = None
        font = FONT_BODY
        if is_today:
            # 今日剛公告 → 整列鮮綠 (蓋過 score 色, 最醒目)
            fill = FILL_TODAY
            font = FONT_TODAY
        elif isinstance(score, (int, float)):
            if score >= 8:
                fill = FILL_HIGHLIGHT  # 鮮黃: 高度超預期
                font = FONT_HIGHLIGHT
            elif score >= 4:
                fill = FILL_HOT  # 暖橘: 有亮點
            elif score <= -4:
                fill = FILL_COLD  # 冷藍: 衰退警示
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=c)
            if fill:
                cell.fill = fill
            cell.font = font
            cell.alignment = ALIGN_LEFT if c in (2, 15) else ALIGN_CENTER
            cell.border = BORDER
        # 公告日期欄獨立色帶 (蓋過分數色, 讓同日視覺成群) — 但今日列已整列鮮綠, 不再重蓋
        if date_col_idx is not None and not is_today:
            d_cell = ws.cell(row=row_idx, column=date_col_idx)
            d_val = d_cell.value
            band = date_to_band.get(d_val)
            if band:
                d_cell.fill = band
                # 維持字型粗體可讀
                d_cell.font = Font(name='Microsoft JhengHei', bold=True, size=10, color='1F3864')

    # 凍結代號 + 名稱 (前 2 欄) + 標題列
    ws.freeze_panes = 'C3'


def write_revenue(ws, monthly_data: list, title: str = '💰 月營收'):
    """月營收分頁 — 對齊 v3 視覺"""
    headers = ['代號', '名稱', '市場', '月份', '當月營收(千元)', 'YoY %', 'MoM %',
               '累計營收(千元)', '累計 YoY %']
    _set_widths(ws, [9, 14, 8, 11, 16, 11, 11, 16, 12])

    _draw_title(ws, 1, title, len(headers), FILL_TITLE_BLUE, big=True)

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER
    ws.row_dimensions[2].height = 22

    for r in monthly_data:
        rev = r.get('revenue')
        rev_k = round(rev / 1000) if rev else None
        accum = r.get('accumulated')
        accum_k = round(accum / 1000) if accum else None
        yoy = r.get('yoy')
        yoy_str = f'{yoy:+.1f}%' if yoy is not None else '—'
        mom = r.get('mom')
        mom_str = f'{mom:+.1f}%' if mom is not None else '—'
        a_yoy = r.get('accum_yoy')
        a_yoy_str = f'{a_yoy:+.1f}%' if a_yoy is not None else '—'
        ws.append([
            r.get('stock_id'),
            r.get('name', ''),
            r.get('market', ''),
            r.get('ym', ''),
            rev_k,
            yoy_str,
            mom_str,
            accum_k,
            a_yoy_str,
        ])

    # 上色: YoY ≥ +50% 鮮黃, ≤ -30% 冷藍
    for row_idx in range(3, ws.max_row + 1):
        cell_yoy = ws.cell(row=row_idx, column=6).value
        fill = None
        font = FONT_BODY
        try:
            yoy_v = float(str(cell_yoy).replace('%', '').replace('+', ''))
            if yoy_v >= 50:
                fill = FILL_HIGHLIGHT
                font = FONT_HIGHLIGHT
            elif yoy_v <= -30:
                fill = FILL_COLD
        except (ValueError, TypeError):
            pass
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=c)
            if fill:
                cell.fill = fill
            cell.font = font
            cell.alignment = ALIGN_LEFT if c == 2 else ALIGN_CENTER
            cell.border = BORDER
        # 千元欄位用千分位
        for col_idx in (5, 8):
            ws.cell(row=row_idx, column=col_idx).number_format = '#,##0'

    ws.freeze_panes = 'C3'


def _style_data_cell(cell, align=None, border=True, font=None):
    cell.font = font or FONT_BODY
    cell.alignment = align or ALIGN_CENTER
    if border:
        cell.border = BORDER


def write_won_full_year(ws, releases: list, q_label: str,
                        first_seen_map: dict = None,
                        today_str: str = None):
    """已贏全年確認名單 — 累計達成率 ≥ 100% 或 虧轉盈 (對應 v3「Q1贏全年確認名單」)

    排序: 評分 desc (主要), latest_eps desc (tiebreaker)
    若有 first_seen_map → 加「公告日期」欄, 今日公告用鮮綠 (蓋過鮮黃 score 色)
    today_str: 'YYYY-MM-DD'. 未傳則用 first_seen_map 最大日期.
    """
    prior_year = _prior_year(q_label)
    rows = [r for r in releases
            if r.get('latest_quarter') == q_label and _is_won_full_year(r)]
    # 評分降冪 (None 視為 -99 排最後), tiebreaker: latest_eps
    rows.sort(key=lambda x: (-(x.get('score') if x.get('score') is not None else -99),
                             -(x.get('latest_eps') or 0)))

    has_date = bool(first_seen_map)
    headers = ['代號', '名稱', f'{q_label} EPS', f'{prior_year}{q_label[-2:]} EPS',
               f'{prior_year} 全年 EPS', '累計/全年', '倍數', '評分']
    widths = [9, 14, 13, 14, 14, 12, 11, 9]
    if has_date:
        headers.append('公告日期')
        widths.append(13)
    _set_widths(ws, widths)

    # 紅色色塊大標題
    title = f'🏆 已確認 {q_label} EPS > {prior_year} 全年 EPS ({len(rows)} 檔)'
    _draw_title(ws, 1, title, len(headers), FILL_TITLE_RED, big=True)
    ws.cell(row=2, column=1, value='一季賺贏去年全年的明星標的（含去年虧損 → 今年轉盈）')
    ws.cell(row=2, column=1).font = Font(name='Microsoft JhengHei', italic=True, size=10, color='808080')
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))

    # Header
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER
    ws.row_dimensions[3].height = 22

    # Data rows (鮮黃高亮 + 公告日期同日色帶)
    for r in rows:
        ach = r.get('achievement_pct')
        pf = r.get('prior_year_full')
        if isinstance(ach, (int, float)):
            ratio_str = f'{ach*100:.1f}%'
            multi_str = f'{ach:.2f}x'
        else:
            ratio_str = '虧轉盈'
            multi_str = '虧轉盈'
        row_data = [
            r.get('stock_id'),
            r.get('name', ''),
            r.get('latest_eps'),
            r.get('yoy_eps'),
            pf,
            ratio_str,
            multi_str,
            r.get('score'),
        ]
        if has_date:
            sid = r.get('stock_id')
            row_data.append(first_seen_map.get(sid, '—'))
        ws.append(row_data)

    # 推導「今日」: 優先用 today_str 參數, 否則用 first_seen_map 最大日期
    today_for_highlight = today_str
    if not today_for_highlight and first_seen_map:
        valid = [d for d in first_seen_map.values() if d and d != '—']
        if valid:
            today_for_highlight = max(valid)

    # 為每個獨特日期分配一個色帶 (同日同色, 不同日交替), 用較淡的色相區分
    # 但「今日」這格特別用鮮綠突顯, 蓋過鮮黃
    date_palette = [
        FILL_HIGHLIGHT,                              # 鮮黃 (非今日的次新日期)
        PatternFill('solid', start_color='FFEB9C'),  # 淡黃
        PatternFill('solid', start_color='FCE4D6'),  # 淡橘
        PatternFill('solid', start_color='E2EFDA'),  # 淡綠
        PatternFill('solid', start_color='DDEBF7'),  # 淡藍
        PatternFill('solid', start_color='F4E1F2'),  # 淡紫
    ]
    date_to_fill: dict = {}
    if has_date:
        unique_dates = []
        for r in rows:
            d = first_seen_map.get(r.get('stock_id'), '—')
            if d not in unique_dates:
                unique_dates.append(d)
        # 不包含今日的其他日期 → 套淡色 palette (新到舊)
        non_today = sorted([d for d in unique_dates if d != today_for_highlight], reverse=True)
        for i, d in enumerate(non_today):
            date_to_fill[d] = date_palette[i % len(date_palette)]
        # 今日 → 鮮綠 (覆蓋, 確保最醒目)
        if today_for_highlight:
            date_to_fill[today_for_highlight] = FILL_TODAY

    for row_idx, r in enumerate(rows, start=4):
        sid = r.get('stock_id')
        first_date = first_seen_map.get(sid, '—') if has_date else None
        is_today = (first_date == today_for_highlight) if today_for_highlight else False
        row_fill = date_to_fill.get(first_date, FILL_HIGHLIGHT) if has_date else FILL_HIGHLIGHT
        row_font = FONT_TODAY if is_today else FONT_HIGHLIGHT
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=c)
            cell.fill = row_fill
            cell.font = row_font
            cell.alignment = ALIGN_CENTER if c != 2 else ALIGN_LEFT
            cell.border = BORDER

    ws.freeze_panes = 'A4'


def write_won_full_year_candidates(ws, releases: list, q_label: str):
    """贏全年候選 — 上季虧損 + 本期轉正 (對應 v3「Q1贏全年候選」)"""
    rows = [r for r in releases
            if r.get('latest_quarter') == q_label
            and r.get('prev_quarter_eps') is not None and r['prev_quarter_eps'] < 0
            and r.get('latest_eps') is not None and r['latest_eps'] > 0
            and not _is_won_full_year(r)]
    rows.sort(key=lambda x: -((x.get('latest_eps') or 0) - (x.get('prev_quarter_eps') or 0)))

    headers = ['代號', '名稱', '評分', f'{q_label} EPS', '上季 EPS', 'QoQ Δ',
               'GM%', 'OPM%', '判斷邏輯']
    _set_widths(ws, [9, 14, 8, 13, 12, 11, 10, 10, 36])

    # 金色色塊大標題 (候選 = 待驗證 = 黃)
    title = f'📋 候選名單：上季虧損 + {q_label} 轉正（{len(rows)} 檔，待全年驗證）'
    _draw_title(ws, 1, title, len(headers), FILL_TITLE_GOLD, big=True)
    ws.cell(row=2, column=1,
            value='說明：上季 EPS 為負、本期已轉正。若全年加總仍虧損 / 微利則本期可能贏全年。')
    ws.cell(row=2, column=1).font = Font(name='Microsoft JhengHei', italic=True, size=10, color='808080')
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER
    ws.row_dimensions[3].height = 22

    for r in rows:
        prev = r.get('prev_quarter_eps')
        latest = r.get('latest_eps') or 0
        qoq_delta = round(latest - (prev or 0), 2)
        ws.append([
            r.get('stock_id'),
            r.get('name', ''),
            r.get('score'),
            latest,
            prev,
            qoq_delta,
            f'{(r.get("gm_pct") or 0)*100:.1f}%' if r.get('gm_pct') is not None else '',
            f'{(r.get("opm_pct") or 0)*100:.1f}%' if r.get('opm_pct') is not None else '',
            f'{r.get("prev_quarter") or "上季"} 虧損 → {q_label} 轉正',
        ])

    # 高分上色 + body 樣式
    for row_idx in range(4, ws.max_row + 1):
        sc = ws.cell(row=row_idx, column=3).value
        is_hot = isinstance(sc, (int, float)) and sc >= 6
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=c)
            if is_hot:
                cell.fill = FILL_HIGHLIGHT
                cell.font = FONT_HIGHLIGHT
            else:
                cell.font = FONT_BODY
            cell.alignment = ALIGN_LEFT if c in (2, 9) else ALIGN_CENTER
            cell.border = BORDER

    ws.freeze_panes = 'A4'


def write_top_eps(ws, releases: list, q_label: str, top_n: int = 30):
    """高 EPS 排行 (對應 v3「Q1高EPS排行」)"""
    rows = [r for r in releases
            if r.get('latest_quarter') == q_label and r.get('latest_eps') is not None]
    rows.sort(key=lambda x: -x['latest_eps'])
    rows = rows[:top_n]

    headers = ['排名', '代號', '名稱', '評分', f'{q_label} EPS', '上季 EPS', 'QoQ Δ',
               '去年全年 EPS', '達成率']
    _set_widths(ws, [7, 9, 14, 8, 13, 12, 11, 14, 11])

    title = f'💰 {q_label} EPS 高低排行（前 {len(rows)} 名）'
    _draw_title(ws, 1, title, len(headers), FILL_TITLE_BLUE, big=True)
    ws.cell(row=2, column=1, value='依當期 EPS 由高到低；達成率 ≥ 100% 整列鮮黃高亮（已贏全年）')
    ws.cell(row=2, column=1).font = Font(name='Microsoft JhengHei', italic=True, size=10, color='808080')
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER
    ws.row_dimensions[3].height = 22

    for i, r in enumerate(rows, 1):
        prev = r.get('prev_quarter_eps')
        latest = r.get('latest_eps')
        qoq_delta = round(latest - prev, 2) if (prev is not None and latest is not None) else None
        ach = r.get('achievement_pct')
        if isinstance(ach, (int, float)):
            ach_str = f'{ach*100:.0f}%'
        elif ach == 'prior_loss':
            ach_str = '虧轉盈'
        else:
            ach_str = '—'
        ws.append([
            i,
            r.get('stock_id'),
            r.get('name', ''),
            r.get('score'),
            latest,
            prev,
            qoq_delta,
            r.get('prior_year_full'),
            ach_str,
        ])

    # 上色: 達成率 ≥ 100% 整列鮮黃高亮
    for row_idx in range(4, ws.max_row + 1):
        ach_cell = ws.cell(row=row_idx, column=9).value
        is_hot = False
        if ach_cell == '虧轉盈':
            is_hot = True
        elif isinstance(ach_cell, str) and ach_cell.endswith('%'):
            try:
                if float(ach_cell.replace('%', '')) >= 100:
                    is_hot = True
            except ValueError:
                pass
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=c)
            if is_hot:
                cell.fill = FILL_HIGHLIGHT
                cell.font = FONT_HIGHLIGHT
            else:
                cell.font = FONT_BODY
            cell.alignment = ALIGN_LEFT if c == 3 else ALIGN_CENTER
            cell.border = BORDER

    ws.freeze_panes = 'A4'


def _detect_q_label(releases: list) -> str:
    """從 releases 中找最常見的 latest_quarter (即「當期」)"""
    counter = {}
    for r in releases:
        q = r.get('latest_quarter')
        if q:
            counter[q] = counter.get(q, 0) + 1
    if not counter:
        return ''
    return max(counter.items(), key=lambda kv: kv[1])[0]


def build_report(date_str: str, releases: list, monthly: list, stats: dict, out_path: str,
                 q_label: str = None, first_seen_map: dict = None,
                 today_str: str = None):
    """產出完整報告

    分頁順序 (v3 風格):
      1. 摘要 — 已贏全年清單 + 觀察重點
      2. 今日新公告 — 完整欄位 + 公告日期色帶
      3. 已贏全年確認名單 — 一季賺贏全年的明星標的 (含公告日期色帶)
      4. 贏全年候選 — 上季虧損 + 本期轉正 (待全年確認)
      5. 高 EPS 排行 — 當期 EPS 前 30
      6. 高分排行 — score ≥ 4 (含公告日期色帶)
      7. 衰退警示 — score ≤ -4
      8. 完整 EPS — 全市場
      9. 月營收

    first_seen_map: {stock_id: 'YYYY-MM-DD'}, 來自 main.py 掃 snapshots。
    """
    if not q_label:
        q_label = _detect_q_label(releases)

    wb = Workbook()
    # 摘要 (含已贏全年清單 + 觀察重點)
    ws_sum = wb.active
    ws_sum.title = '摘要'
    write_summary(ws_sum, date_str, stats, releases=releases, q_label=q_label)

    # 今日新公告 (評分降冪 + latest_eps tiebreaker)
    new_only = sorted([r for r in releases if r.get('is_new')],
                      key=lambda x: (-(x.get('score') if x.get('score') is not None else -99),
                                     -(x.get('latest_eps') or 0)))
    # 今日參考: 優先 today_str 參數, 否則退到 date_str
    today_for_highlight = today_str or date_str

    if new_only:
        ws_new = wb.create_sheet('今日新公告')
        write_releases(ws_new, new_only, f'🆕 今日新公告 ({len(new_only)} 檔)', FILL_TITLE_BLUE,
                       first_seen_map=first_seen_map, today_str=today_for_highlight)

    # v3 風格三個 sheet — 都依 q_label 篩選
    if q_label:
        won = [r for r in releases
               if r.get('latest_quarter') == q_label and _is_won_full_year(r)]
        if won:
            ws_won = wb.create_sheet(f'{q_label}贏全年確認')
            write_won_full_year(ws_won, releases, q_label, first_seen_map=first_seen_map,
                                today_str=today_for_highlight)

        cand = [r for r in releases
                if r.get('latest_quarter') == q_label
                and r.get('prev_quarter_eps') is not None and r['prev_quarter_eps'] < 0
                and (r.get('latest_eps') or 0) > 0
                and not _is_won_full_year(r)]
        if cand:
            ws_cand = wb.create_sheet(f'{q_label}贏全年候選')
            write_won_full_year_candidates(ws_cand, releases, q_label)

        cur_q = [r for r in releases if r.get('latest_quarter') == q_label
                 and r.get('latest_eps') is not None]
        if cur_q:
            ws_top = wb.create_sheet(f'{q_label}高EPS排行')
            write_top_eps(ws_top, releases, q_label, top_n=30)

    # 高分排行 (score ≥ 4) — tiebreaker latest_eps desc
    hot = sorted([r for r in releases if (r.get('score') or 0) >= 4],
                 key=lambda x: (-(x.get('score') or 0), -(x.get('latest_eps') or 0)))
    if hot:
        ws_hot = wb.create_sheet('高分排行')
        write_releases(ws_hot, hot, f'🔥 高分排行 (評分 ≥ 4，{len(hot)} 檔)', FILL_TITLE_RED,
                       first_seen_map=first_seen_map, today_str=today_for_highlight)

    # 衰退警示 (score ≤ -4)
    cold = sorted([r for r in releases if (r.get('score') or 0) <= -4],
                  key=lambda x: x.get('score') or 0)
    if cold:
        ws_cold = wb.create_sheet('衰退警示')
        write_releases(ws_cold, cold, f'⚠️ 衰退警示 (評分 ≤ -4，{len(cold)} 檔)', FILL_TITLE_GOLD,
                       first_seen_map=first_seen_map, today_str=today_for_highlight)

    # 完整 EPS
    full = sorted(releases,
                  key=lambda x: (-(x.get('score') if x.get('score') is not None else -99),
                                 -(x.get('latest_eps') or 0)))
    ws_full = wb.create_sheet('完整 EPS')
    write_releases(ws_full, full, f'📊 完整 EPS 全市場 ({len(full)} 檔)', FILL_TITLE_BLUE,
                   first_seen_map=first_seen_map, today_str=today_for_highlight)

    # 月營收
    if monthly:
        ws_rev = wb.create_sheet('月營收')
        write_revenue(ws_rev, monthly)

    # 移除無印表機設定避免開檔卡
    from openpyxl.worksheet.page import PageMargins, PrintOptions, PrintPageSetup
    for sname in wb.sheetnames:
        ws = wb[sname]
        new_setup = PrintPageSetup(worksheet=ws)
        new_setup.paperSize = 9
        new_setup.orientation = 'landscape'
        ws.page_setup = new_setup
        ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)
        ws.print_options = PrintOptions()
        ws.print_area = None

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


if __name__ == '__main__':
    # 自測
    import sys
    sys.stdout.reconfigure(encoding='utf-8')
    sample_releases = [
        {
            'stock_id': '5386', 'name': '青雲', 'market': '上市',
            'is_new': True,
            'latest_quarter': '2026Q1', 'latest_eps': 43.05, 'yoy_eps': 1.27,
            'yoy': {'pct': 32.9, 'delta': 41.78},
            'qoq': {'pct': 6.71, 'delta': 37.47},
            'accumulated': {'value': 43.05, 'quarters_count': 1},
            'prior_year_full': 8.78, 'achievement_pct': 4.9,
            'gm_pct': 0.27, 'opm_pct': 0.23, 'nonop_pct': -0.02,
            'score': 9, 'level': '高度超預期', 'label': '🔥 +9',
            'reasons': ['QoQ 爆發 +671%', 'YoY 爆發 +3290%', '已賺贏去年全年 (490%)'],
        },
    ]
    out = build_report('2026-05-06', sample_releases, [],
                       {'new_count': 1, 'hot_count': 1, 'total_count': 1},
                       'reports/test_report.xlsx')
    print(f'產出: {out}')
